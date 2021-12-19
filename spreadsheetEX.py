from openpyxl import Workbook
import openpyxl
wb = openpyxl.load_workbook(filename= "inventory.xlsx")
product_list = wb["Sheet1"]

suppliersSUM = {}
lowInventory = {}

AAAProducts = 0
BBBProducts = 0
CCCProducts = 0



for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = int(product_list.cell(product_row, 2).value)
    product = int(product_list.cell(product_row, 1).value)
    price = float(product_list.cell(product_row, 3).value)

    if supplier_name == "AAA Company":
        AAAProducts+=1
    if supplier_name == "BBB Company":
        BBBProducts+=1
    if supplier_name == "CCC Company":
        CCCProducts+=1
    
    
    if inventory < 10:
        lowInventory[product] = inventory

        
suppliersSUM = {"AAA": AAAProducts,  "BBB": BBBProducts , "CCC": CCCProducts}
print(suppliersSUM)
print(lowInventory)
